import pytest
import openpyxl


class BaseClass():
    
    def getrowcount(file,sheetName):
        book = openpyxl.load_workbook(file)
        sheet=book[sheetName]
        return (sheet.max_row)
    
    def getcolumncount(file,sheetName):
        book = openpyxl.load_workbook(file)
        sheet=book[sheetName]
        return (sheet.max_column)
    
    def readData(file,sheetName,rownum,columnno):
        book = openpyxl.load_workbook(file)
        sheet=book[sheetName]
        return sheet.cell(row=rownum, column=columnno).value
    
    def writedata(file,sheetName,rownum,columnno,data):
        book = openpyxl.load_workbook(file)
        sheet=book[sheetName]
        sheet.cell(row=rownum, column=columnno).value=data
        book.save(file)
    