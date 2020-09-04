import openpyxl
class data():
    @staticmethod
    def getdata(test_case_name):
        book = openpyxl.load_workbook("/home/pi/Documents/project_Guru99/user credential.xlsx")
        sheet=book.active
        Dict={}
        for i in range (1,sheet.max_row+1):
            if sheet.cell(row=i,column=1) == test_case_name:
                for j in range (1,sheet.max_column+1):                
                    Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
        return[Dict]